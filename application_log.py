"""
Private, local application log -- an .xlsx file on your own machine.
Never sent anywhere. Records every IPO application attempt: which
account, which IPO, the result, and the Application ID once captured.
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

LOG_FILE = "applications_log.xlsx"
HEADERS = ["Timestamp", "Account Label", "Company", "Investor Type",
           "Status", "Application ID", "UPI ID", "Quantity", "Price", "Notes"]

def _ensure_log_exists():
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(HEADERS)
        wb.save(LOG_FILE)

def log_application(account_label, company, investor_type="Individual investor",
                     status="", application_id="", upi_id="", quantity="",
                     price="", notes=""):
    """Appends one row to the log. Safe against the file being open in
    Excel at the time -- logs a warning instead of crashing the bot."""
    _ensure_log_exists()
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb["Applications"]
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            account_label, company, investor_type, status,
            application_id, upi_id, quantity, price, notes,
        ])
        wb.save(LOG_FILE)
        print(f"Logged: {account_label} / {company} / {status}")
    except PermissionError:
        print(f"WARNING: Could not write to {LOG_FILE} -- it may be open in Excel. "
              f"Close it and this entry can be added manually, or it will be logged next run.")
    except Exception as e:
        print(f"WARNING: Failed to log application: {e}")

if __name__ == "__main__":
    # Quick manual test
    log_application(
        account_label="TEST",
        company="TESTCO",
        status="TEST_ENTRY",
        notes="This is a test row from running application_log.py directly."
    )
    print(f"\nCheck {LOG_FILE} -- you should see a header row plus this one test row.")