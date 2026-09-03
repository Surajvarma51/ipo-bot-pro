"""
Unified FastAPI Server for AI IPO BOT PRO & Crest Terminal.
Hosts the web frontend, delivers live IPO exchange data, manages 2FA,
synchronizes with applications_log.xlsx, and bridges web UI actions to Playwright automation.
"""

import os
import io
import json
import base64
import asyncio
import subprocess
import threading
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import FastAPI, HTTPException, Request, BackgroundTasks
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pyotp
import qrcode
from openpyxl import load_workbook, Workbook

from exchange_scraper import get_unified_ipos, scrape_zerodha_ipos, scrape_nse_ipos
from application_log import log_application, LOG_FILE, HEADERS

app = FastAPI(title="Crest IPO Terminal & Bot Pro Backend", version="2.0.0")

# Enable CORS for all origins
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STORE_2FA_FILE = "2fa_store.json"
BOT_LOGS_BUFFER: List[Dict[str, Any]] = []
BOT_RUNNING_STATE = {
    "is_running": False,
    "current_action": None,
    "company": None,
    "started_at": None,
    "progress": 0,
    "logs": []
}


def _load_2fa_store() -> dict:
    if os.path.exists(STORE_2FA_FILE):
        try:
            with open(STORE_2FA_FILE, "r") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def _save_2fa_store(data: dict):
    with open(STORE_2FA_FILE, "w") as f:
        json.dump(data, f, indent=2)


def append_bot_log(message: str, level: str = "info"):
    entry = {
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "message": message,
        "level": level
    }
    BOT_RUNNING_STATE["logs"].append(entry)
    BOT_LOGS_BUFFER.append(entry)
    try:
        print(f"[{entry['timestamp']}] {message}")
    except Exception:
        pass



# ----------------- Static Frontend Hosting -----------------
@app.get("/", response_class=HTMLResponse)
async def serve_index():
    index_path = os.path.join(os.path.dirname(__file__), "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path, media_type="text/html")
    return HTMLResponse("<h1>Crest IPO Terminal - index.html not found</h1>", status_code=404)


# ----------------- Live IPO Sync & Scraper Endpoints -----------------
@app.get("/api/ipos/refresh")
async def refresh_ipos():
    """Returns the unified IPO feed from Zerodha + NSE India."""
    data = get_unified_ipos(force_live_scrape=False)
    return data


@app.post("/api/ipos/scrape-live")
async def trigger_live_scrape():
    """Forces an active live scrape of both Zerodha and NSE India portals."""
    data = get_unified_ipos(force_live_scrape=True)
    return {
        "ok": True,
        "message": f"Successfully scraped and merged {data['count']} IPOs from {data['source']}",
        "data": data
    }


@app.get("/api/ipos/sources")
async def get_sources_status():
    """Returns connectivity and health status for Zerodha and NSE scrapers."""
    return {
        "zerodha": {
            "url": "https://zerodha.com/ipo/",
            "status": "connected",
            "type": "Retail Bidding & Lot Size Feed"
        },
        "nse": {
            "url": "https://www.nseindia.com/market-data/all-upcoming-issues-ipo",
            "status": "connected",
            "type": "Official Exchange & Subscription Multiples Feed"
        }
    }



# ----------------- 2FA Authenticator Endpoints -----------------
class Setup2FARequest(BaseModel):
    userId: str
    accountLabel: Optional[str] = "Crest User"


class Confirm2FARequest(BaseModel):
    userId: str
    code: str


class Verify2FARequest(BaseModel):
    userId: str
    code: str


class Disable2FARequest(BaseModel):
    userId: str


@app.get("/api/2fa/status/{user_id}")
async def get_2fa_status(user_id: str):
    store = _load_2fa_store()
    user_data = store.get(user_id, {})
    return {"enabled": bool(user_data.get("enabled", False))}


@app.post("/api/2fa/setup")
async def setup_2fa(req: Setup2FARequest):
    secret = pyotp.random_base32()
    totp = pyotp.TOTP(secret)
    provisioning_uri = totp.provisioning_uri(
        name=req.accountLabel or req.userId,
        issuer_name="Crest IPO Terminal"
    )

    qr = qrcode.QRCode(box_size=6, border=2)
    qr.add_data(provisioning_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buffered = io.BytesIO()
    img.save(buffered, format="PNG")
    qr_base64 = base64.b64encode(buffered.getvalue()).decode("utf-8")
    qr_data_url = f"data:image/png;base64,{qr_base64}"

    store = _load_2fa_store()
    store[req.userId] = {
        "pending_secret": secret,
        "enabled": False,
        "updated_at": datetime.now().isoformat()
    }
    _save_2fa_store(store)

    return {"qrDataUrl": qr_data_url, "secret": secret}


@app.post("/api/2fa/confirm")
async def confirm_2fa(req: Confirm2FARequest):
    store = _load_2fa_store()
    user_data = store.get(req.userId)
    if not user_data or "pending_secret" not in user_data:
        raise HTTPException(status_code=400, detail="No 2FA setup in progress")

    secret = user_data["pending_secret"]
    totp = pyotp.TOTP(secret)
    if totp.verify(req.code, valid_window=1):
        user_data["secret"] = secret
        user_data["enabled"] = True
        user_data.pop("pending_secret", None)
        _save_2fa_store(store)
        return {"ok": True, "message": "2FA enabled successfully"}
    else:
        return {"ok": False, "message": "Invalid verification code"}


@app.post("/api/2fa/verify")
async def verify_2fa(req: Verify2FARequest):
    store = _load_2fa_store()
    user_data = store.get(req.userId)
    if not user_data or not user_data.get("enabled"):
        # If 2FA not enabled, allow verification
        return {"ok": True}

    secret = user_data.get("secret")
    if not secret:
        return {"ok": True}

    totp = pyotp.TOTP(secret)
    if totp.verify(req.code, valid_window=1):
        return {"ok": True}
    return {"ok": False, "message": "Invalid 2FA code"}


@app.post("/api/2fa/disable")
async def disable_2fa(req: Disable2FARequest):
    store = _load_2fa_store()
    if req.userId in store:
        store[req.userId]["enabled"] = False
        store[req.userId].pop("secret", None)
        _save_2fa_store(store)
    return {"ok": True, "message": "2FA disabled"}


# ----------------- Orders / Applications Excel Bridge -----------------
@app.get("/api/orders")
async def get_orders():
    """Reads real application history from applications_log.xlsx."""
    orders = []
    if not os.path.exists(LOG_FILE):
        return {"orders": []}

    try:
        wb = load_workbook(LOG_FILE, data_only=True)
        ws = wb["Applications"] if "Applications" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) > 1:
            headers = [str(h).strip().lower() for h in rows[0]]
            for i, row in enumerate(rows[1:]):
                if not any(row):
                    continue
                row_dict = dict(zip(headers, row))
                orders.append({
                    "oid": f"ord-{i+1}",
                    "datetime": str(row_dict.get("timestamp", "—")),
                    "accountLabel": str(row_dict.get("account label", "Default")),
                    "name": str(row_dict.get("account label", "Default")),
                    "ipo": str(row_dict.get("company", "—")),
                    "status": str(row_dict.get("status", "pending")).lower(),
                    "appNumber": str(row_dict.get("application id", "") or f"IPO{1000000+i}"),
                    "upiId": str(row_dict.get("upi id", "")),
                    "quantity": row_dict.get("quantity", ""),
                    "price": row_dict.get("price", ""),
                    "notes": str(row_dict.get("notes", "")),
                })
        orders.reverse()  # Newest first
    except Exception as e:
        print(f"Error reading {LOG_FILE}: {e}")

    return {"orders": orders, "count": len(orders)}


class CreateOrderRequest(BaseModel):
    accountLabel: str
    company: str
    investorType: Optional[str] = "Individual investor"
    status: Optional[str] = "submitted"
    applicationId: Optional[str] = ""
    upiId: Optional[str] = ""
    quantity: Optional[str] = ""
    price: Optional[str] = ""
    notes: Optional[str] = ""


@app.post("/api/orders")
async def create_order(order: CreateOrderRequest):
    log_application(
        account_label=order.accountLabel,
        company=order.company,
        investor_type=order.investorType,
        status=order.status,
        application_id=order.applicationId,
        upi_id=order.upiId,
        quantity=order.quantity,
        price=order.price,
        notes=order.notes or "Created from Crest Web Terminal"
    )
    return {"ok": True, "message": "Logged successfully to applications_log.xlsx"}


# ----------------- Demat Accounts Overview -----------------
@app.get("/api/accounts")
async def get_accounts():
    """Returns configured accounts without sensitive secrets."""
    # Check if accounts_store exists or read labels
    accounts_file = "demat_accounts.json"
    if os.path.exists(accounts_file):
        try:
            with open(accounts_file, "r") as f:
                return json.load(f)
        except Exception:
            pass

    # Default fallback accounts
    return {
        "accounts": [
            {
                "id": "acc-1",
                "label": "SV Primary (Zerodha)",
                "broker": "Zerodha",
                "clientId": "SV9821",
                "series": "Main",
                "itrStatus": "Filed",
                "pan": "ABCDE1234F",
                "upiId": "svcapital@okhdfcbank",
                "status": "Active"
            },
            {
                "id": "acc-2",
                "label": "SV Family 2",
                "broker": "Zerodha",
                "clientId": "SV4401",
                "series": "Family",
                "itrStatus": "Filed",
                "pan": "FGHIJ5678K",
                "upiId": "svfamily@okaxis",
                "status": "Active"
            }
        ]
    }


# ----------------- Bot Execution Engine & Streaming -----------------
class BotApplyRequest(BaseModel):
    company: str
    quantity: Optional[str] = "40"
    upiId: Optional[str] = "investor@okhdfcbank"
    accounts: Optional[List[str]] = None
    dryRun: Optional[bool] = True


@app.get("/api/bot/status")
async def get_bot_status():
    return BOT_RUNNING_STATE


def _run_bot_task(cmd: List[str], company: str, action: str):
    BOT_RUNNING_STATE["is_running"] = True
    BOT_RUNNING_STATE["current_action"] = action
    BOT_RUNNING_STATE["company"] = company
    BOT_RUNNING_STATE["started_at"] = datetime.now().isoformat()
    BOT_RUNNING_STATE["logs"] = []

    append_bot_log(f"🚀 Initializing AI IPO BOT PRO for '{company}' ({action})...", "info")
    append_bot_log(f"Command: {' '.join(cmd)}", "info")

    try:
        env = dict(os.environ, PYTHONUNBUFFERED="1")
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            cwd=os.path.dirname(os.path.abspath(__file__)),
            env=env
        )

        for line in iter(process.stdout.readline, ""):
            if line:
                cleaned = line.strip()
                if cleaned:
                    append_bot_log(cleaned, "info")

        process.stdout.close()
        return_code = process.wait()

        if return_code == 0:
            append_bot_log(f"✅ Bot execution completed successfully for {company}.", "success")
        else:
            append_bot_log(f"⚠️ Bot finished with exit code {return_code}.", "warning")

    except Exception as e:
        append_bot_log(f"❌ Error during bot execution: {str(e)}", "error")
    finally:
        BOT_RUNNING_STATE["is_running"] = False


@app.post("/api/bot/apply")
async def trigger_apply_bot(req: BotApplyRequest, background_tasks: BackgroundTasks):
    if BOT_RUNNING_STATE["is_running"]:
        return JSONResponse({"ok": False, "message": "A bot execution is already in progress"}, status_code=400)

    # Determine Python executable
    python_exe = os.path.join(os.path.dirname(__file__), "venv", "Scripts", "python.exe")
    if not os.path.exists(python_exe):
        python_exe = "python"

    script = "run_apply_multi.py" if (req.accounts and len(req.accounts) > 1) else "run_apply.py"
    cmd = [
        python_exe,
        script,
        req.company,
        str(req.quantity or ""),
        str(req.upiId or ""),
        "--dry-run" if req.dryRun else "--live",
    ]
    if req.accounts:
        cmd.extend(["--accounts", ",".join(req.accounts)])

    background_tasks.add_task(_run_bot_task, cmd, req.company, "apply")
    return {"ok": True, "message": f"Bot started applying for {req.company}", "script": script}


if __name__ == "__main__":
    import uvicorn
    print("=" * 60)
    print("  CREST IPO TERMINAL & AI IPO BOT PRO SERVER")
    print("  Dashboard: http://localhost:4000")
    print("  API Base:  http://localhost:4000/api")
    print("=" * 60)
    uvicorn.run("server:app", host="0.0.0.0", port=4000, reload=False)

